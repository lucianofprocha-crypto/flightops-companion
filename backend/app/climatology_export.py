"""
Exportação dos dados de Airport Intelligence (climatologia + eventos
abaixo dos mínimos) em Excel (.xlsx) e PDF — fecha os itens "Exportar
Excel" / "Exportar PDF" do V1.0 original (ver docs/VISAO_E_ROADMAP.md).

Recebe os mesmos dicts já calculados por climatology.compute_climatology()
e minima_events.build_events_summary() — não busca nem reprocessa nada,
só formata o que o app já tem em mãos.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

# Paleta VFR/MVFR/IFR/LIFR — mesma usada no frontend (CATEGORY_COLORS em
# app.js), pra manter a leitura visual consistente entre tela e export.
_CATEGORY_HEX = {
    "VFR": "22C55E",
    "MVFR": "3B82F6",
    "IFR": "F59E0B",
    "LIFR": "EF4444",
    "UNKNOWN": "6B7280",
}
_STATUS_HEX = {"green": "22C55E", "yellow": "F59E0B", "red": "EF4444"}

_INK_HEX = "1A2333"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def _header_row(ws, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_INK_HEX)
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _category_fill(cat: str) -> PatternFill:
    return PatternFill("solid", fgColor=_CATEGORY_HEX.get(cat, "6B7280"))


def build_climatology_excel(climatology: dict, events: dict | None = None) -> bytes:
    wb = Workbook()

    # --- Resumo -------------------------------------------------------
    ws = wb.active
    ws.title = "Resumo"
    icao = climatology.get("icao", "?")
    period = climatology.get("period", "?")
    ws["A1"] = f"FlightOps Companion — Airport Intelligence — {icao}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Período: {period} · Gerado em {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%MZ')}"
    ws["A2"].font = Font(color="666666")

    sample = climatology.get("sample", {})
    temp = climatology.get("temperature_c", {})
    wind = climatology.get("wind_kt", {})
    vis = climatology.get("visibility_m", {})

    rows = [
        ("Observações na amostra", sample.get("observations")),
        ("Período coberto", f"{(sample.get('start') or '')[:10]} a {(sample.get('end') or '')[:10]}"),
        ("Temperatura média (°C)", temp.get("mean")),
        ("Temperatura mín/máx (°C)", f"{temp.get('min')} / {temp.get('max')}"),
        ("Vento médio (kt)", wind.get("mean_speed")),
        ("Rajada máxima (kt)", wind.get("max_gust")),
        ("Direção dominante", wind.get("dominant_direction")),
        ("Visibilidade média (m)", vis.get("mean")),
        ("Visibilidade mínima (m)", vis.get("min")),
        ("% condições adversas (IFR/LIFR)", climatology.get("adverse_conditions_pct")),
    ]
    if events:
        headline = events.get("headline", {})
        rows += [
            ("— Eventos abaixo dos mínimos —", ""),
            ("Disponibilidade (acima dos mínimos)", f"{headline.get('availability_pct')}%"),
            ("Horas abaixo dos mínimos", headline.get("hours_below_minima")),
            ("Nº de eventos", headline.get("event_count")),
            ("Maior evento", headline.get("longest_event_label")),
            ("Menor visibilidade (m)", headline.get("min_visibility_m")),
            ("Menor teto (ft)", headline.get("min_ceiling_ft")),
        ]

    r = 4
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Categoria de voo").font = Font(bold=True)
    _header_row(ws, r, ["Categoria", "Contagem", "%"])
    r += 1
    for cat, data in climatology.get("flight_category", {}).items():
        ws.cell(row=r, column=1, value=cat).fill = _category_fill(cat)
        ws.cell(row=r, column=2, value=data.get("count"))
        ws.cell(row=r, column=3, value=data.get("pct"))
        r += 1

    _autosize(ws, [34, 22, 10])

    # --- Por mês --------------------------------------------------------
    ws2 = wb.create_sheet("Por mês")
    _header_row(ws2, 1, ["Mês", "Observações", "% condições adversas", "Temp. média (°C)"])
    for i, (month_key, data) in enumerate(sorted(climatology.get("by_month", {}).items(), key=lambda x: int(x[0])), start=2):
        ws2.cell(row=i, column=1, value=MESES[int(month_key) - 1])
        ws2.cell(row=i, column=2, value=data.get("count"))
        ws2.cell(row=i, column=3, value=data.get("adverse_pct"))
        ws2.cell(row=i, column=4, value=data.get("mean_temp_c"))
    _autosize(ws2, [10, 14, 20, 18])

    # --- Por hora ---------------------------------------------------------
    ws3 = wb.create_sheet("Por hora (UTC)")
    _header_row(ws3, 1, ["Hora UTC", "Observações", "% condições adversas", "Temp. média (°C)"])
    for i, (hour_key, data) in enumerate(sorted(climatology.get("by_hour_utc", {}).items(), key=lambda x: int(x[0])), start=2):
        ws3.cell(row=i, column=1, value=f"{int(hour_key):02d}h")
        ws3.cell(row=i, column=2, value=data.get("count"))
        ws3.cell(row=i, column=3, value=data.get("adverse_pct"))
        ws3.cell(row=i, column=4, value=data.get("mean_temp_c"))
    _autosize(ws3, [10, 14, 20, 18])

    if events:
        # --- Heatmap mês x hora -------------------------------------------
        ws4 = wb.create_sheet("Heatmap")
        ws4.cell(row=1, column=1, value="Hora \\ Mês").font = Font(bold=True)
        for col, mes in enumerate(MESES, start=2):
            c = ws4.cell(row=1, column=col, value=mes)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=_INK_HEX)
        heatmap = events.get("heatmap", {})
        for hour in range(24):
            row = hour + 2
            ws4.cell(row=row, column=1, value=f"{hour:02d}h").font = Font(bold=True)
            for month in range(1, 13):
                cell_data = heatmap.get(str(month), {}).get(str(hour), {})
                pct = cell_data.get("below_minima_pct", 0)
                ws4.cell(row=row, column=month + 1, value=pct)
        pct_range = f"B2:M25"
        ws4.conditional_formatting.add(
            pct_range,
            ColorScaleRule(
                start_type="num", start_value=0, start_color="FFFFFF",
                end_type="num", end_value=100, end_color="EF4444",
            ),
        )
        _autosize(ws4, [10] + [7] * 12)

        # --- Calendário -----------------------------------------------------
        ws5 = wb.create_sheet("Calendário")
        _header_row(ws5, 1, ["Data", "Status", "Minutos abaixo dos mínimos", "Observações no dia"])
        for i, (date_key, info) in enumerate(sorted(events.get("calendar", {}).items()), start=2):
            ws5.cell(row=i, column=1, value=date_key)
            status_cell = ws5.cell(row=i, column=2, value=info.get("status"))
            status_cell.fill = PatternFill("solid", fgColor=_STATUS_HEX.get(info.get("status"), "FFFFFF"))
            ws5.cell(row=i, column=3, value=info.get("minutes_below_minima"))
            ws5.cell(row=i, column=4, value=info.get("observation_count"))
        _autosize(ws5, [14, 10, 26, 18])

        # --- Eventos ----------------------------------------------------
        ws6 = wb.create_sheet("Eventos")
        _header_row(
            ws6, 1,
            ["Data", "Início (UTC)", "Fim (UTC)", "Duração", "Causa", "Pior categoria",
             "Menor vis. (m)", "Menor teto (ft)", "Fim incerto?"],
        )
        for i, ev in enumerate(events.get("events", []), start=2):
            start_dt = ev["start"]
            end_dt = ev["end"]
            ws6.cell(row=i, column=1, value=start_dt[:10])
            ws6.cell(row=i, column=2, value=start_dt[11:16])
            ws6.cell(row=i, column=3, value=end_dt[11:16])
            ws6.cell(row=i, column=4, value=ev.get("duration_label"))
            ws6.cell(row=i, column=5, value=ev.get("cause") or "—")
            cat_cell = ws6.cell(row=i, column=6, value=ev.get("worst_category"))
            cat_cell.fill = _category_fill(ev.get("worst_category"))
            ws6.cell(row=i, column=7, value=ev.get("min_visibility_m"))
            ws6.cell(row=i, column=8, value=ev.get("min_ceiling_ft"))
            ws6.cell(row=i, column=9, value="Sim" if ev.get("end_uncertain") else "")
        _autosize(ws6, [12, 12, 12, 10, 10, 14, 14, 14, 12])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def _fmt(value, suffix: str = "") -> str:
    """Formata um valor pra exibição, evitando o literal 'None' no PDF
    quando o dado não está disponível na amostra (ex: sem rajada registrada)."""
    return f"{value}{suffix}" if value is not None else "—"


def _styles():
    ss = getSampleStyleSheet()
    ss.add(ParagraphStyle(name="H1FO", parent=ss["Heading1"], fontSize=17, textColor=colors.HexColor(f"#{_INK_HEX}"), spaceAfter=2))
    ss.add(ParagraphStyle(name="H2FO", parent=ss["Heading2"], fontSize=12.5, textColor=colors.HexColor(f"#{_INK_HEX}"), spaceBefore=12, spaceAfter=5))
    ss.add(ParagraphStyle(name="Body9", parent=ss["Normal"], fontSize=9, leading=12))
    ss.add(ParagraphStyle(name="Small", parent=ss["Normal"], fontSize=7.5, textColor=colors.HexColor("#666666"), leading=10))
    ss.add(ParagraphStyle(name="Cell", parent=ss["Normal"], fontSize=7.5, leading=9.5))
    return ss


def _table(data, col_widths, header_bg=True):
    t = Table(data, colWidths=col_widths, repeatRows=1)
    cmds = [
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    if header_bg:
        cmds += [
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_INK_HEX}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ]
    t.setStyle(TableStyle(cmds))
    return t


def build_climatology_pdf(climatology: dict, events: dict | None = None) -> bytes:
    styles = _styles()
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=18 * mm, bottomMargin=16 * mm, leftMargin=16 * mm, rightMargin=16 * mm,
        title="FlightOps Companion - Airport Intelligence",
    )

    icao = climatology.get("icao", "?")
    period = climatology.get("period", "?")
    flow: list = []
    flow.append(Paragraph(f"FlightOps Companion &mdash; Airport Intelligence &mdash; {icao}", styles["H1FO"]))
    now = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%MZ")
    flow.append(Paragraph(f"Período: {period} &middot; Gerado em {now}", styles["Small"]))
    flow.append(HRFlowable(width="100%", thickness=0.75, color=colors.HexColor(f"#{_INK_HEX}"), spaceBefore=6, spaceAfter=8))

    # Resumo
    sample = climatology.get("sample", {})
    temp = climatology.get("temperature_c", {})
    wind = climatology.get("wind_kt", {})
    vis = climatology.get("visibility_m", {})
    flow.append(Paragraph("Resumo", styles["H2FO"]))
    resumo_linhas = [
        f"Amostra: {_fmt(sample.get('observations'))} observações · {(sample.get('start') or '')[:10]} a {(sample.get('end') or '')[:10]}",
        f"Temperatura: média {_fmt(temp.get('mean'), '°C')} (mín {_fmt(temp.get('min'), '°')} / máx {_fmt(temp.get('max'), '°')})",
        f"Vento: médio {_fmt(wind.get('mean_speed'), ' kt')} · rajada máx. {_fmt(wind.get('max_gust'), ' kt')} · direção dominante {wind.get('dominant_direction') or '—'}",
        f"Visibilidade: média {_fmt(vis.get('mean'), ' m')} · mínima {_fmt(vis.get('min'), ' m')}",
        f"Condições adversas (IFR/LIFR): {_fmt(climatology.get('adverse_conditions_pct'), '%')}",
    ]
    for line in resumo_linhas:
        flow.append(Paragraph(line, styles["Body9"]))
    flow.append(Spacer(1, 6))

    # Categorias de voo
    flow.append(Paragraph("Categorias de voo", styles["H2FO"]))
    cat_data = [["Categoria", "Contagem", "%"]]
    cat_colors_by_row = []
    for cat, data in climatology.get("flight_category", {}).items():
        cat_data.append([cat, str(data.get("count")), f"{data.get('pct')}%"])
        cat_colors_by_row.append(_CATEGORY_HEX.get(cat, "6B7280"))
    t = _table(cat_data, [40 * mm, 30 * mm, 20 * mm])
    extra = TableStyle(
        [("BACKGROUND", (0, i + 1), (0, i + 1), colors.HexColor(f"#{c}")) for i, c in enumerate(cat_colors_by_row)]
        + [("TEXTCOLOR", (0, i + 1), (0, i + 1), colors.white) for i in range(len(cat_colors_by_row))]
        + [("FONTNAME", (0, i + 1), (0, i + 1), "Helvetica-Bold") for i in range(len(cat_colors_by_row))]
    )
    t.setStyle(extra)
    flow.append(t)
    flow.append(Spacer(1, 8))

    # Por mês
    flow.append(Paragraph("Por mês", styles["H2FO"]))
    mes_data = [["Mês", "Observações", "% adversas", "Temp. média"]]
    for month_key, data in sorted(climatology.get("by_month", {}).items(), key=lambda x: int(x[0])):
        mes_data.append([
            MESES[int(month_key) - 1], str(data.get("count")),
            f"{data.get('adverse_pct')}%", f"{data.get('mean_temp_c')}°C" if data.get("mean_temp_c") is not None else "—",
        ])
    flow.append(_table(mes_data, [25 * mm, 30 * mm, 28 * mm, 28 * mm]))
    flow.append(Spacer(1, 8))

    # Por hora
    flow.append(Paragraph("Por hora (UTC)", styles["H2FO"]))
    hora_data = [["Hora", "Observações", "% adversas", "Temp. média"]]
    for hour_key, data in sorted(climatology.get("by_hour_utc", {}).items(), key=lambda x: int(x[0])):
        hora_data.append([
            f"{int(hour_key):02d}h", str(data.get("count")),
            f"{data.get('adverse_pct')}%", f"{data.get('mean_temp_c')}°C" if data.get("mean_temp_c") is not None else "—",
        ])
    flow.append(_table(hora_data, [25 * mm, 30 * mm, 28 * mm, 28 * mm]))

    if events:
        headline = events.get("headline", {})
        flow.append(Spacer(1, 10))
        flow.append(Paragraph("Eventos abaixo dos mínimos (IFR/LIFR)", styles["H2FO"]))
        flow.append(
            Paragraph(
                f"Disponibilidade: {_fmt(headline.get('availability_pct'), '%')} &middot; "
                f"{_fmt(headline.get('hours_below_minima'))}h abaixo dos mínimos em {_fmt(headline.get('event_count'))} evento(s) "
                f"&middot; maior evento: {headline.get('longest_event_label') or '—'} &middot; "
                f"menor vis.: {_fmt(headline.get('min_visibility_m'), ' m')} &middot; menor teto: {_fmt(headline.get('min_ceiling_ft'), ' ft')}",
                styles["Body9"],
            )
        )
        flow.append(Spacer(1, 4))

        ev_rows = [["Data", "Início", "Fim", "Duração", "Causa", "Pior cat."]]
        for ev in events.get("events", []):
            ev_rows.append([
                ev["start"][:10], ev["start"][11:16], ev["end"][11:16] + (" (?)" if ev.get("end_uncertain") else ""),
                ev.get("duration_label"), ev.get("cause") or "—", ev.get("worst_category"),
            ])
        if len(ev_rows) > 1:
            t_ev = _table(ev_rows, [22 * mm, 18 * mm, 20 * mm, 20 * mm, 24 * mm, 22 * mm])
            cat_style_cmds = []
            for i, ev in enumerate(events.get("events", []), start=1):
                hexcolor = _CATEGORY_HEX.get(ev.get("worst_category"), "6B7280")
                cat_style_cmds.append(("TEXTCOLOR", (5, i), (5, i), colors.HexColor(f"#{hexcolor}")))
                cat_style_cmds.append(("FONTNAME", (5, i), (5, i), "Helvetica-Bold"))
            t_ev.setStyle(TableStyle(cat_style_cmds))
            flow.append(t_ev)
        else:
            flow.append(Paragraph("Nenhum evento neste período.", styles["Body9"]))

    flow.append(Spacer(1, 14))
    flow.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc"), spaceBefore=2, spaceAfter=6))
    flow.append(
        Paragraph(
            "Projeção estatística a partir do histórico observado (IEM Mesonet) — não é uma previsão numérica (NWP). "
            "Gerado automaticamente pelo FlightOps Companion.",
            styles["Small"],
        )
    )

    doc.build(flow)
    return buf.getvalue()
